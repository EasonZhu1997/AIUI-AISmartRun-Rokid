#!/usr/bin/env python3
# 生成《AISmartRun 商业化任务总表》—— 双产品线(AIUI 眼镜端 + APK 生态)全部待办,
# 按功能象限(商业影响 × 投入成本)组织。数据源:双项目盘点 + 商业化差距清单 + 90 天路线。
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Microsoft YaHei"
INK = "1F2937"          # 深墨(正文)
HEAD_BG = "0F5132"      # 深绿表头
HEAD_FG = "FFFFFF"
BORDER = Border(*[Side(style="thin", color="D0D5DD")] * 4)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 象限配色(商业影响 × 投入成本)
QUAD = {
    "Q1 快赢":   ("C6EFCE", "高影响 · 低投入 → 先做"),
    "Q2 战略":   ("BDD7EE", "高影响 · 高投入 → 立项排期"),
    "Q3 顺手":   ("FFF2CC", "低影响 · 低投入 → 有空补"),
    "Q4 缓做":   ("E9D5FF", "低影响 · 高投入 → 触发式/暂缓"),
}
STATUS_FILL = {
    "已完成":   "D1FAE5",
    "进行中":   "FEF3C7",
    "未开始":   "FFFFFF",
    "待外部":   "FFE4CC",
}
PROJ_FILL = {
    "AIUI眼镜端": "E0F2FE",
    "APK手机端":  "EDE9FE",
    "Garmin表":   "FEF9C3",
    "后端":       "F1F5F9",
    "共用":       "FCE7F3",
}

# ── 任务数据 ──────────────────────────────────────────────────
# (ID, 项目, 功能分组, 任务, 象限, 影响, 投入, 状态, 排期, 工期, 阻塞/依赖, 备注/证据)
TASKS = [
    # 支付收入
    ("P01","后端","支付收入","定价三档拍板(Free / Pro / Pro+AR)","Q1 快赢","高","低","未开始","M2","0.5天","商业模式选型","PRICING.md 仅草案,非上架承诺"),
    ("P02","APK手机端","支付收入","Apple IAP 接入(海外首发收款通道)","Q2 战略","高","高","未开始","M2","3-5天","依赖 iOS 过审","无任何 IAP 端点;海外不依赖国内主体,优先于微信"),
    ("P03","后端","支付收入","微信支付 v3 验签 + 商户号 + 订单表","Q2 战略","高","高","未开始","M3+","5天+","公司主体+ICP","subscription.py 硬闸门 _WECHAT..=False 禁用中"),
    ("P04","后端","支付收入","订阅表 / require_pro / 402 额度墙 / 邀请送 PRO","Q1 快赢","高","中","已完成","已交付","—","—","唯一真实 PRO 发放路径=邀请裂变"),
    # 账号打通
    ("A01","AIUI眼镜端","账号打通","修复 anon-login 契约断裂(后端必填 app_key)","Q1 快赢","高","低","已完成","本轮","—","—","coach_app_key 注入+无 key 跳过;记忆链路曾静默全灭"),
    ("A02","AIUI眼镜端","账号打通","眼镜跑步数据落后端 runs 表(source=aiui)","Q1 快赢","高","低","已完成","本轮","—","—","lib/run_upload.js+待传队列+14 测试;后端零改动"),
    ("A03","后端","账号打通","匿名 device_id ↔ 手机号账号绑定端点","Q2 战略","高","中","未开始","M1","3天","眼镜扫码绑定","口径2026-07-08:眼镜端暂不做登录,先跑通硬件MVP;绑定推迟到硬件验证后"),
    ("A04","后端","账号打通","短信登录生产接线(腾讯签名/模板)","Q1 快赢","高","低","进行中","M0-M1","1天","腾讯审核","代码就绪 provider 未接;PRODUCTION_CHECKLIST P0"),
    # 留存
    ("R01","AIUI眼镜端","留存","眼镜跑后总结卡(单屏摘要)","Q1 快赢","高","低","未开始","M0-M1","1-2天","复用 session 快照","跑者需求矩阵判为上架前必做;跑完即消失是致命缺口"),
    ("R02","AIUI眼镜端","留存","自定义最大心率(年龄/手动,修 Z5 播报准确)","Q1 快赢","中","低","未开始","M1","0.5天","—","maxHr 现固定 190,对年长用户 Z5 迟报;PRD P1"),
    ("R03","AIUI眼镜端","留存","RSC/FTMS 外设接入 HUD 优先数据源","Q3 顺手","中","低","未开始","M1","1天","—","解析库已完成,只差接线;提配速/距离精度"),
    ("R04","后端","留存","远程推送(FCM/APNs/微信模板)","Q2 战略","高","中","未开始","M1","3天","—","notify.py 自述 thin SHELL,主动触达能力=0"),
    ("R05","后端","留存","D1/D3/D7 拉活 job","Q3 顺手","中","低","进行中","M1","0.5天","依赖 R04 推送","已写,无通道静默跳过"),
    ("R06","后端","留存","周报/连跑天数/成就/本地提醒","Q3 顺手","中","中","已完成","已交付","—","—","PRODUCTION_CHECKLIST P1 已勾"),
    # 合规
    ("C01","后端","合规","密钥/密码全量轮换(root/API key 已入 git 史)","Q1 快赢","高","低","未开始","M0","0.5天","—","唯一'正在流血'风险;RUNBOOK.md/deploy_coach.py 明文"),
    ("C02","后端","合规","生产环境切 prod(关万能验证码/mock-pay)","Q1 快赢","高","低","未开始","M0","0.5天","C01 先行","线上仍 ENVIRONMENT=dev,123456 万能码可用"),
    ("C03","后端","合规","品牌域名 + 正确 TLS(现 nip.io 证书不匹配)","Q1 快赢","高","低","未开始","M0","0.5天","需注册域名","play_readiness_audit --strict 唯一 blocker"),
    ("C04","后端","合规","修生产 nginx /coach-svc POST body 转发 bug","Q1 快赢","高","低","未开始","M0","0.5天","当面确认后改","不修则匿名登录/记忆/记录生产全挂"),
    ("C05","共用","合规","法务文本定稿(主体/邮箱/管辖 + 律师审)","Q3 顺手","中","中","进行中","M1","2天","—","docs/launch/01,02 待填项"),
    ("C06","APK手机端","合规","隐私链路(同意门/导出/PIPL 注销)","Q3 顺手","中","中","已完成","已交付","—","—","ConsentGate + test_privacy_delete 背书"),
    ("C07","后端","合规","国内 ICP 备案 / 软著 / 生成式 AI 登记","Q4 缓做","中","高","未开始","M2+","—","需大陆部署","服务器在新加坡,国内收费另需公司主体"),
    ("C08","Garmin表","合规","SimHei 字体授权处置(换思源黑体重烤)","Q3 顺手","低","低","未开始","M1","0.5天","CIQ 提审前置","04-GARMIN_CIQ §2.2 版权风险"),
    ("C09","共用","合规","健康数据免责落到用户可见面","Q3 顺手","中","低","进行中","M0-M1","0.5天","—","心率=PIPL 敏感数据;运动建议免责边界"),
    # 分发上架
    ("D01","AIUI眼镜端","分发上架","真机 Alpha 全量回归(A-01~A-14)","Q1 快赢","高","中","未开始","M0","1-2天","设备已备,硬阻塞","必过 A-02/03/08/10/11 + P0≥10/12"),
    ("D02","AIUI眼镜端","分发上架","AIUI Studio 正式签名 + 打包 + 提审","Q1 快赢","高","低","未开始","M0-M1","0.5-1天","依赖 D01","部分依赖 Rokid 官方文档"),
    ("D03","AIUI眼镜端","分发上架","Agent 补 accelerometer 权限重新打包上传","Q1 快赢","高","低","未开始","M0","0.5天","—","否则真机 IMU 计步/估距不可用"),
    ("D04","APK手机端","分发上架","iOS App Store 提审","Q2 战略","高","中","待外部","M0被动","—","苹果审核","2026-07-07 已提交 1.0.0 build1,等 App Review"),
    ("D05","APK手机端","分发上架","Google Play 开发者账号注册($25)","Q1 快赢","高","低","未开始","M0","0.5天","启动硬时钟","越晚注册越晚上架"),
    ("D06","APK手机端","分发上架","Google Play 12 人 × 14 天封闭测试","Q2 战略","高","高","未开始","M0-M1","14天时钟","依赖 D05","不可压缩;个人账号发布前置"),
    ("D07","APK手机端","分发上架","Play 商店资料 + Data Safety + 内容分级","Q2 战略","中","中","未开始","M1","1天","—","03-GOOGLE_PLAY §4/§5"),
    ("D08","Garmin表","分发上架","Garmin Connect IQ 提审(watch-app)","Q3 顺手","低","中","未开始","M1","2天","依赖 C08","免费零外部依赖;7 项清单勾 1"),
    ("D09","AIUI眼镜端","分发上架","AIUI Agent Store 提审(眼镜第 4 条线)","Q1 快赢","高","低","未开始","M1","0.5天","依赖 D01/D02","Agent ID 已有"),
    ("D10","共用","分发上架","品牌名统一 = AISmartRun(域名/邮箱/法务替换)","Q1 快赢","中","低","进行中","M0","0.5天","需买域名","三端显示名已一致;funpizza 字样待清"),
    ("D11","共用","分发上架","商店素材(真机截图 + 高清图标)","Q3 顺手","中","低","进行中","M0-M1","1天","依赖真机","Play 真机截图缺;iOS 截图齐"),
    # 基础设施
    ("I01","后端","基础设施","云端对象存储备份(替换本机拉取镜像)","Q3 顺手","中","低","进行中","M1","1天","—","现为过渡态,strict 审计已过"),
    ("I02","后端","基础设施","SQLite → Postgres + pgvector 迁移","Q4 缓做","中","高","未开始","M2","3天","DAU>100 触发","条件式规划,非当前阻塞"),
    ("I03","后端","基础设施","外部告警 / TSDB(替换 admin 人工日检)","Q4 缓做","低","中","未开始","M2","2天","—","现靠 /admin/metrics + 人工"),
    ("I04","后端","基础设施","修 /health 被 voice 服务占用","Q3 顺手","低","低","未开始","M0","0.1天","—","coach 深健康检查未从公网暴露"),
    # 内容运营
    ("N01","后端","内容运营","路线库 23 条 + 7 预置训练计划","Q3 顺手","中","中","已完成","已交付","—","—","content_lib + plan_presets 底子已有"),
    ("N02","后端","内容运营","批量生成 50 路线 + 10 计划模板 + 审核流","Q4 缓做","中","高","未开始","M2","2天+","—","ROADMAP P1-⑧;等上架后留存供给"),
]

COLS = ["ID","项目","功能分组","任务","象限","商业影响","投入成本","状态","排期","工期估","阻塞/依赖","备注 / 证据"]
WIDTHS = [6, 11, 11, 34, 9, 8, 8, 8, 8, 9, 15, 40]

wb = Workbook()

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color=HEAD_FG, size=10)
        cell.fill = PatternFill("solid", fgColor=HEAD_BG)
        cell.alignment = CENTER
        cell.border = BORDER

def title_block(ws, title, subtitle, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(name=FONT, bold=True, size=16, color="0F5132")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(name=FONT, size=9, color="6B7280", italic=True)
    s.alignment = Alignment(horizontal="left", vertical="center")

# ── Sheet 1: 说明 & 图例 ──────────────────────────────────────
ws0 = wb.active
ws0.title = "说明"
title_block(ws0, "AISmartRun 商业化任务总表",
            "双产品线(AIUI 眼镜端 + APK 生态)· 按功能象限(商业影响 × 投入成本)组织 · 数据源:双项目盘点 + 商业化差距清单 + 90 天路线",
            6)
r = 4
ws0.cell(row=r, column=1, value="一、功能象限定义(两轴:纵=商业影响,横=投入成本)").font = Font(name=FONT, bold=True, size=11, color=INK)
r += 1
# 2x2 矩阵示意
matrix = [
    ["", "投入低", "投入高"],
    ["影响高", "Q1 快赢 · 先做", "Q2 战略 · 立项排期"],
    ["影响低", "Q3 顺手 · 有空补", "Q4 缓做 · 触发式/暂缓"],
]
for i, rowvals in enumerate(matrix):
    for j, v in enumerate(rowvals):
        cell = ws0.cell(row=r + i, column=1 + j, value=v)
        cell.alignment = CENTER
        cell.border = BORDER
        cell.font = Font(name=FONT, bold=(i == 0 or j == 0), size=10, color=INK)
        if v.startswith("Q1"): cell.fill = PatternFill("solid", fgColor=QUAD["Q1 快赢"][0])
        elif v.startswith("Q2"): cell.fill = PatternFill("solid", fgColor=QUAD["Q2 战略"][0])
        elif v.startswith("Q3"): cell.fill = PatternFill("solid", fgColor=QUAD["Q3 顺手"][0])
        elif v.startswith("Q4"): cell.fill = PatternFill("solid", fgColor=QUAD["Q4 缓做"][0])
        elif i == 0 or j == 0: cell.fill = PatternFill("solid", fgColor="E5E7EB")
        ws0.row_dimensions[r + i].height = 24
r += len(matrix) + 1
ws0.cell(row=r, column=1, value="二、图例").font = Font(name=FONT, bold=True, size=11, color=INK)
r += 1
legend = [
    ("排期", "M0 = 第 1-4 周(止血+启动硬时钟) | M1 = 第 5-8 周(汇流+免费增长) | M2 = 第 9-13 周(开收款) | M3+ = 之后/国内线"),
    ("状态", "已完成 / 进行中 / 未开始 / 待外部(等审核方)"),
    ("象限", "Q1 快赢 | Q2 战略 | Q3 顺手 | Q4 缓做(颜色见上矩阵)"),
    ("项目", "AIUI眼镜端 | APK手机端 | Garmin表 | 后端 | 共用"),
]
for k, v in legend:
    kc = ws0.cell(row=r, column=1, value=k); kc.font = Font(name=FONT, bold=True, size=10, color=INK); kc.alignment = LEFT
    ws0.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    vc = ws0.cell(row=r, column=2, value=v); vc.font = Font(name=FONT, size=10, color=INK); vc.alignment = LEFT
    r += 1
r += 1
ws0.cell(row=r, column=1, value="三、实时统计(公式引用《总表》,数据变自动更新)").font = Font(name=FONT, bold=True, size=11, color=INK)
r += 1
# 统计块用 COUNTIF 公式引用总表(总表列: 象限=E, 状态=H, 排期=I, 项目=B)
stat_header_row = r
stats = [
    ("象限分布", [("Q1 快赢","E"),("Q2 战略","E"),("Q3 顺手","E"),("Q4 缓做","E")]),
    ("状态分布", [("已完成","H"),("进行中","H"),("未开始","H"),("待外部","H")]),
    ("排期分布", [("M0","I"),("M1","I"),("M2","I"),("M3+","I")]),
]
col = 1
for name, items in stats:
    hc = ws0.cell(row=r, column=col, value=name)
    hc.font = Font(name=FONT, bold=True, size=10, color=HEAD_FG); hc.fill = PatternFill("solid", fgColor=HEAD_BG)
    hc.alignment = CENTER; hc.border = BORDER
    ws0.cell(row=r, column=col+1, value="数").font = Font(name=FONT, bold=True, size=10, color=HEAD_FG)
    ws0.cell(row=r, column=col+1).fill = PatternFill("solid", fgColor=HEAD_BG)
    ws0.cell(row=r, column=col+1).alignment = CENTER; ws0.cell(row=r, column=col+1).border = BORDER
    # 统计值在生成时算好写成数值:静态交付表无需依赖 Excel 重算,任何查看器都正确显示、零公式错误。
    col_index = {"B": 1, "E": 4, "H": 7, "I": 8}
    for i, (label, colletter) in enumerate(items):
        lc = ws0.cell(row=r+1+i, column=col, value=label)
        lc.font = Font(name=FONT, size=10, color=INK); lc.alignment = LEFT; lc.border = BORDER
        if colletter == "E" and label in QUAD:
            lc.fill = PatternFill("solid", fgColor=QUAD[label][0])
        count = sum(1 for t in TASKS if t[col_index[colletter]] == label)
        fc = ws0.cell(row=r+1+i, column=col+1, value=count)
        fc.font = Font(name=FONT, size=10, color="000000"); fc.alignment = CENTER; fc.border = BORDER
    col += 3
# 合计
r2 = r + 5
ws0.cell(row=r2, column=1, value="任务总数").font = Font(name=FONT, bold=True, size=10, color=INK)
tc = ws0.cell(row=r2, column=2, value=len(TASKS))
tc.font = Font(name=FONT, bold=True, size=10, color="000000")
done_n = sum(1 for t in TASKS if t[7] == "已完成")
ws0.cell(row=r2, column=3, value="已完成").font = Font(name=FONT, size=10, color="6B7280")
ws0.cell(row=r2, column=4, value=f"{done_n} / {len(TASKS)}").font = Font(name=FONT, bold=True, size=10, color="0F5132")
for w, wd in zip("ABCDEF", [12,12,4,12,4,12]):
    ws0.column_dimensions[w].width = wd
ws0.sheet_view.showGridLines = False

# ── Sheet 2: 总表 ─────────────────────────────────────────────
ws = wb.create_sheet("总表")
title_block(ws, "总表 · 全部待办", "共 %d 项 · 可用表头筛选/排序 · 冻结首列与表头" % len(TASKS), len(COLS))
hr = 4
for j, name in enumerate(COLS, 1):
    ws.cell(row=hr, column=j, value=name)
style_header(ws, hr, len(COLS))
# 象限 → 排序键(Q1..Q4),让总表按象限聚拢
qorder = {"Q1 快赢":1, "Q2 战略":2, "Q3 顺手":3, "Q4 缓做":4}
rows = sorted(TASKS, key=lambda t: (qorder[t[4]], t[0]))
for i, t in enumerate(rows):
    rr = hr + 1 + i
    for j, val in enumerate(t, 1):
        cell = ws.cell(row=rr, column=j, value=val)
        cell.font = Font(name=FONT, size=10, color=INK)
        cell.border = BORDER
        cell.alignment = LEFT if j in (4, 11, 12) else CENTER
    # 象限着色(列 5)
    ws.cell(row=rr, column=5).fill = PatternFill("solid", fgColor=QUAD[t[4]][0])
    ws.cell(row=rr, column=5).font = Font(name=FONT, size=10, bold=True, color=INK)
    # 状态着色(列 8)
    stf = STATUS_FILL.get(t[7])
    if stf: ws.cell(row=rr, column=8).fill = PatternFill("solid", fgColor=stf)
    if t[7] == "已完成":
        ws.cell(row=rr, column=8).font = Font(name=FONT, size=10, bold=True, color="0F5132")
    # 项目着色(列 2)
    pf = PROJ_FILL.get(t[1])
    if pf: ws.cell(row=rr, column=2).fill = PatternFill("solid", fgColor=pf)
    # 影响/投入 高亮"高"
    for cc in (6, 7):
        if ws.cell(row=rr, column=cc).value == "高":
            ws.cell(row=rr, column=cc).font = Font(name=FONT, size=10, bold=True, color="B91C1C")
    ws.row_dimensions[rr].height = 30
for j, wd in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(j)].width = wd
ws.freeze_panes = "B5"
ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(COLS))}{hr + len(rows)}"
ws.sheet_view.showGridLines = False

# ── Sheet 3: 功能象限矩阵(2x2,ID 分格) ──────────────────────
wsq = wb.create_sheet("功能象限矩阵")
title_block(wsq, "功能象限矩阵", "每格列出任务 ID(去掉已完成),先清 Q1,再排 Q2,Q3 见缝插针,Q4 触发式", 4)
def cell_ids(quad):
    ids = [f"{t[0]} {t[3][:14]}" for t in TASKS if t[4] == quad and t[7] != "已完成"]
    return "\n".join(ids) if ids else "(无)"
grid = [
    (4, 2, "影响高", "E5E7EB", None),
    (4, 3, "投入低 ↓", "E5E7EB", None),
    (4, 4, "投入高 ↓", "E5E7EB", None),
    (5, 1, "影响高 →", "E5E7EB", None),
    (5, 3, "Q1 快赢", QUAD["Q1 快赢"][0], cell_ids("Q1 快赢")),
    (5, 4, "Q2 战略", QUAD["Q2 战略"][0], cell_ids("Q2 战略")),
    (6, 1, "影响低 →", "E5E7EB", None),
    (6, 3, "Q3 顺手", QUAD["Q3 顺手"][0], cell_ids("Q3 顺手")),
    (6, 4, "Q4 缓做", QUAD["Q4 缓做"][0], cell_ids("Q4 缓做")),
]
for (rr, cc, label, fill, body) in grid:
    cell = wsq.cell(row=rr, column=cc)
    if body is None:
        cell.value = label
        cell.font = Font(name=FONT, bold=True, size=11, color=INK)
    else:
        cell.value = f"{label}\n{QUAD[label][1]}\n\n{body}"
        cell.font = Font(name=FONT, size=9, color=INK)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = BORDER
wsq.column_dimensions["A"].width = 10
for c in "BCD":
    wsq.column_dimensions[c].width = 40
wsq.column_dimensions["B"].width = 8
wsq.row_dimensions[5].height = 150
wsq.row_dimensions[6].height = 150
wsq.sheet_view.showGridLines = False

# ── Sheet 4: 90 天路线(M0/M1/M2/M3+) ─────────────────────────
wsr = wb.create_sheet("90天路线")
title_block(wsr, "90 天商业化路线", "按排期分组 · 利用双项目互补(APK 已有跑后管线,AIUI 已在本轮打通数据落库)", 6)
RCOLS = ["排期", "ID", "项目", "任务", "工期估", "阻塞/依赖"]
RW = [12, 6, 11, 40, 9, 18]
hr = 4
for j, name in enumerate(RCOLS, 1):
    wsr.cell(row=hr, column=j, value=name)
style_header(wsr, hr, len(RCOLS))
phase_order = {"M0":0,"M0-M1":1,"M0被动":1,"M1":2,"M2":3,"M2+":4,"M3+":5,"本轮":-2,"已交付":-1}
rrows = sorted([t for t in TASKS], key=lambda t: (phase_order.get(t[8], 9), qorder[t[4]], t[0]))
phase_fill = {"M0":"C6EFCE","M0-M1":"D9F2D9","M0被动":"D9F2D9","M1":"BDD7EE","M2":"FFF2CC","M2+":"FDE9C8","M3+":"E9D5FF","本轮":"D1FAE5","已交付":"D1FAE5"}
i = 0
for t in rrows:
    rr = hr + 1 + i
    vals = [t[8], t[0], t[1], t[3], t[9], t[10]]
    for j, val in enumerate(vals, 1):
        cell = wsr.cell(row=rr, column=j, value=val)
        cell.font = Font(name=FONT, size=10, color=INK)
        cell.border = BORDER
        cell.alignment = LEFT if j in (4, 6) else CENTER
    pf = phase_fill.get(t[8])
    if pf: wsr.cell(row=rr, column=1).fill = PatternFill("solid", fgColor=pf)
    if t[7] == "已完成":
        wsr.cell(row=rr, column=4).font = Font(name=FONT, size=10, color="9CA3AF", strike=True)
    wsr.row_dimensions[rr].height = 28
    i += 1
for j, wd in enumerate(RW, 1):
    wsr.column_dimensions[get_column_letter(j)].width = wd
wsr.freeze_panes = "A5"
wsr.sheet_view.showGridLines = False

out = "/Users/apple/Desktop/Project/AIUI_AISmartRun/docs/商业化任务总表.xlsx"
wb.save(out)
print("saved:", out, "| tasks:", len(TASKS))

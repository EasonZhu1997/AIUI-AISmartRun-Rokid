#!/usr/bin/env python3
# 生成《AIUI 眼镜端硬件测试执行表》—— MVP 硬件真机测试为主(用户口径 2026-07-08:
# 眼镜端暂不做登录,后端记忆/上传链路暂缓,只保留其"无 key 本地降级"必测路径)。
# 数据源: docs/ALPHA_TEST_MATRIX.md + DEVICES.md + PRODUCT_PM_REVIEW。
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Microsoft YaHei"
INK = "1F2937"
HEAD_BG = "0F5132"
HEAD_FG = "FFFFFF"
BORDER = Border(*[Side(style="thin", color="D0D5DD")] * 4)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

GROUP_FILL = {
    "①冒烟":     "D1FAE5",
    "②BLE心率":  "DBEAFE",
    "③恢复按键": "FEF3C7",
    "④语音教练": "FCE7F3",
    "⑤降级边界": "E9D5FF",
    "⑥观察项":   "F1F5F9",
}
PRI_FILL = {"P0必过": "FECACA", "P0": "FFE4CC", "P1": "FFF9C4", "暂缓": "E5E7EB"}

# (ID, 组, 优先级, 用例, 前置条件, 执行步骤, 通过标准, 设备)
CASES = [
    # ① 冒烟主线
    ("A-01","①冒烟","P0","首屏进入","Agent 已装(带 accelerometer 权限重传包)","打开 AISmartRun","默认进首页;显示单眼镜就绪+心率就绪状态(已关闭/不可用/已记住/待配对之一);首页不发起蓝牙连接;无错误弹窗","眼镜"),
    ("A-02","①冒烟","P0必过","一键开跑","首页停留","首页点『开跑』或 Enter/Space/GlobalHook","3 秒内进 HUD 并开始计时;首句语音『开跑，呼吸放稳。』","眼镜"),
    ("A-03","①冒烟","P0必过","无心率跑步 10 分钟","不开任何心率设备","戴眼镜跑/走 10 分钟","无心率列;时长持续刷新;步频随动作变化;距离/配速有值;不崩溃","眼镜"),
    ("A-04","①冒烟","P0","加速度计不可用兜底","浏览器 Craft 预览(宿主无 Accelerometer)","Craft 打开 run_hud","显示『仅计时』态,时长走、其余不显示假值;不崩溃","浏览器"),
    # ② BLE 心率(硬件重点)
    ("A-05","②BLE心率","P0","设备页扫描配对","Fenix 8 开广播心率 或 ESP32 已烧 HRS 固件","首页→设备页→搜索设备","可发现标准 HRS 设备;连接验证后保存为首选;状态『已记住』","眼镜+Fenix8/ESP32"),
    ("A-06","②BLE心率","P0","已记住设备优先连接","A-05 已保存首选;设备继续广播","回首页→开跑进 HUD","HUD 自动优先连首选设备;收到有效 bpm 后同屏出现心率列+区间点阵","眼镜+Fenix8/ESP32"),
    ("A-07","②BLE心率","P0","心率断连回退","A-06 心率已接入,跑步中","关闭设备广播(Fenix 关广播/ESP32 断电)","8 秒内静默回单眼镜模式;心率列消失;跑步计时不中断","眼镜+Fenix8/ESP32"),
    ("H-01","②BLE心率","P0","断连自动重连","A-07 后设备重新开播;留在 HUD","重新打开广播,或息屏再回 HUD(onShow 重试)","回到 HUD 触发自动重连;心率列恢复;无重复连接/崩溃","眼镜+ESP32"),
    ("H-02","②BLE心率","P1","首扫手势门槛确认(=B-01)","清除首选设备","冷启动→首页直接开跑,观察 HUD 自动扫描是否被拒","记录:onLoad 自动扫描能否成功;被拒则确认产品口径『首次需设备页授权』成立","眼镜+ESP32"),
    ("H-03","②BLE心率","P1","异常包边界(nRF 模拟)","第二台 Android 装 nRF Connect 建 GATT server","模拟 HR=0/字段缺失/notify 停止不断连","HR=0 或非法包不显示心率;notify 停 8 秒回单眼镜;不崩溃","眼镜+安卓机"),
    ("H-04","②BLE心率","P1","反面设备优雅失败","Apple Watch 不做任何配置","设备页扫描时 Apple Watch 在旁","不误连;无标准 HRS 的设备被忽略或提示不支持;不崩溃","眼镜+AppleWatch"),
    # ③ 恢复与按键
    ("A-08","③恢复按键","P0必过","息屏自动暂停与恢复","跑步中(有/无心率各测一次)","息屏 1 分钟→回到 HUD","暂停期间时长不增加;显示『已暂停』;恢复后 10 秒内步频恢复刷新;无双倍跳秒/重复传感器","眼镜"),
    ("A-13","③恢复按键","P0","防误触双击退出","跑步中","单击 Backspace 等 3 秒;再测 3 秒内双击","单击显示『再按一次结束』,3 秒不按第二次继续跑;双击才结束返回","眼镜"),
    ("A-11","③恢复按键","P0必过","全页面物理 Backspace","逐页进入","首页/设备/设置/HUD/教练页分别按 Backspace","首页退出;设备/设置返回;HUD 走双击确认;教练进行中先取消、空闲则返回;无资源泄漏","眼镜"),
    ("H-05","③恢复按键","P1","方向键+Enter 全链路","—","首页/设备页/设置页用方向键移动焦点+Enter 激活","焦点边框正确移动;Enter/Space/GlobalHook 激活当前项;设置页可改步长/开关","眼镜"),
    # ④ 语音教练(宿主能力,非登录)
    ("A-09","④语音教练","P0","AI 语音问答闭环","跑步中(先跑 2 分钟出数据)","进教练页→按键开始说话→问『现在配速怎么样』","ASR 正确识别;LanguageModel 短答(一句话,含真实配速);TTS 播出或文本可读","眼镜+手机网络"),
    ("A-10","④语音教练","P0必过","断网规则兜底","手机开飞行模式","教练页提问","本地规则回答;有实时数据引用数据,无数据答『暂无运动数据』口径;不编造数字","眼镜"),
    ("H-06","④语音教练","P1","主动语音播报","跑步中持续 5 分钟+","观察整公里/每 5 分钟/Z4/Z5 播报","对应节点播出短句(≤15字);Z5 持续时每分钟重复;不打断数据显示","眼镜+ESP32(拉高心率值)"),
    ("H-07","④语音教练","P1","LLM 挂起超时兜底","弱网(手机限速/信号差)","教练页提问,观察>10 秒无回答场景","10 秒超时后落规则兜底回答;不永久停在『思考』","眼镜"),
    # ⑤ 降级边界(不做登录口径下的必测)
    ("A-12a","⑤降级边界","P0","无 key 跳过登录(本轮口径)","不注入 coach_app_key(默认状态)","教练页完成一次问答","客户端不发登录请求(不付超时);教练回答不受影响;记忆静默关闭","眼镜"),
    ("A-14a","⑤降级边界","P1","跑步摘要入队不上传","不注入 key;跑 2 分钟双击退出","退出后再进首页","无网络请求发出;待传队列保留(cap 5);首页无卡顿","眼镜"),
    ("A-12b","⑤降级边界","暂缓","记忆链路端到端(需后端)","注入 coach_app_key + 修生产 nginx","问答→查后端记录","匿名登录成功;记忆上下文注入;记录写回——**本轮暂缓(不做登录)**","眼镜+后端"),
    ("A-14b","⑤降级边界","暂缓","跑步记录落库补传(需后端)","注入 key","跑步→退出→首页补传→查 GET /api/runs","source=aiui 记录落库,指标一致——**本轮暂缓(不做登录)**","眼镜+后端"),
    # ⑥ 观察项
    ("B-02","⑥观察项","P1","1Hz 功耗与发热","满电开跑","连续跑 20 分钟","记录:电量消耗 %/发热主观分(1-5)/有无卡顿掉帧","眼镜"),
    ("B-03","⑥观察项","P1","IMU 距离误差标定","已知 400m 操场或 1km 直路;对照 Fenix 8","按标准距离跑,对比 HUD 距离","记录误差百分比;>20% 需调步长默认值或文案再降预期","眼镜+Fenix8"),
    ("B-04","⑥观察项","P1","TTS 延迟","A-09 可用","问答 5 次","记录:回答生成→播报开始的延迟均值","眼镜"),
    ("B-05","⑥观察项","P1","英文预览宽度","—","打开英文预览 HTML 对照真机字宽","无溢出;溢出则记录截图","浏览器"),
    ("B-06","⑥观察项","P1","Studio 包与本地包差异","AIUI Studio 可用","官方打包 vs 本地 .aix reader 检查","页面/权限/版本一致;记录差异","PC"),
    ("B-07","⑥观察项","P1","CXR-L 共存验证","手机 side-load relay-apk","同副眼镜:AIUI Agent 与 CXR-L HUD 切换","记录能否共存/互相抢占;结论回写 DUAL_PRODUCT_COMMERCIAL_EVAL 缝3","眼镜+手机"),
]

COLS = ["ID","组","优先级","用例","前置条件","执行步骤","通过标准","设备","结果","实测记录","修复跟进"]
WIDTHS = [7, 10, 9, 16, 22, 26, 34, 14, 8, 22, 18]

wb = Workbook()

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color=HEAD_FG, size=10)
        cell.fill = PatternFill("solid", fgColor=HEAD_BG)
        cell.alignment = CENTER
        cell.border = BORDER

def title_block(ws, title, subtitle, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(name=FONT, bold=True, size=16, color="0F5132")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(name=FONT, size=9, color="6B7280", italic=True)

# ── Sheet1 执行看板 ───────────────────────────────────────────
ws0 = wb.active
ws0.title = "执行看板"
title_block(ws0, "AIUI 眼镜端 · 硬件测试执行表",
            "MVP 口径(2026-07-08):暂不做登录,后端记忆/上传链路暂缓;硬件真机测试为主。结果在《测试用例》页填写,本页统计自动更新",
            6)
r = 4
ws0.cell(row=r, column=1, value="一、退出门槛").font = Font(name=FONT, bold=True, size=11, color=INK)
r += 1
gates = [
    "必过项(5):A-02 一键开跑 / A-03 无心率跑步 / A-08 息屏暂停恢复 / A-10 断网兜底 / A-11 物理按键",
    "P0 用例(A 系 + H-01)至少 10/12 通过(A-12b、A-14b 暂缓项不计入分母)",
    "A-06 不过 → 产品口径改『手动授权后自动优先连接』;A-09 不过但 A-10 过 → 可先做文本教练 Alpha",
    "全部必过项通过 → 真机成熟度可从 88 分复评;随后走 AIUI Studio 签名提审",
]
for g in gates:
    ws0.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws0.cell(row=r, column=1, value="· " + g)
    c.font = Font(name=FONT, size=10, color=INK); c.alignment = LEFT
    ws0.row_dimensions[r].height = 18
    r += 1
r += 1
ws0.cell(row=r, column=1, value="二、进度统计(自动)").font = Font(name=FONT, bold=True, size=11, color=INK)
r += 1
stat_labels = ["通过", "失败", "阻塞", "跳过", "未测"]
for j, lab in enumerate(stat_labels):
    hc = ws0.cell(row=r, column=1 + j, value=lab)
    hc.font = Font(name=FONT, bold=True, size=10, color=HEAD_FG)
    hc.fill = PatternFill("solid", fgColor=HEAD_BG); hc.alignment = CENTER; hc.border = BORDER
for j, lab in enumerate(stat_labels):
    fc = ws0.cell(row=r + 1, column=1 + j, value=f'=COUNTIF(测试用例!I:I,"{lab}")')
    fc.font = Font(name=FONT, size=11, bold=True, color="000000"); fc.alignment = CENTER; fc.border = BORDER
r += 2
ws0.cell(row=r, column=1, value="必过项通过数").font = Font(name=FONT, size=10, color=INK)
fc = ws0.cell(row=r, column=2, value='=COUNTIFS(测试用例!C:C,"P0必过",测试用例!I:I,"通过")&" / 5"')
fc.font = Font(name=FONT, bold=True, size=11, color="0F5132")
r += 2
ws0.cell(row=r, column=1, value="三、建议执行顺序(半天可完成主线)").font = Font(name=FONT, bold=True, size=11, color=INK)
r += 1
order = [
    ("第1批·冒烟(15min)", "A-01 → A-02 → A-03(可缩短到3分钟先验刷新) → A-04(浏览器)"),
    ("第2批·BLE(40min)", "烧录 ESP32 固件 → A-05 → A-06 → A-07 → H-01 → H-02;Fenix 8 复测 A-05~07"),
    ("第3批·恢复按键(20min)", "A-08(有/无心率各一次) → A-13 → A-11 → H-05"),
    ("第4批·语音(20min)", "A-09 → H-06 → H-07 → A-10(飞行模式)"),
    ("第5批·降级(10min)", "A-12a → A-14a(均为默认无 key 状态,顺手完成)"),
    ("外场·观察项", "B-02 功耗 / B-03 距离标定(操场) / B-04 TTS 延迟;B-05~07 室内补"),
]
for name, steps in order:
    nc = ws0.cell(row=r, column=1, value=name)
    nc.font = Font(name=FONT, bold=True, size=10, color=INK); nc.alignment = LEFT; nc.border = BORDER
    nc.fill = PatternFill("solid", fgColor="F0FDF4")
    ws0.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    sc = ws0.cell(row=r, column=2, value=steps)
    sc.font = Font(name=FONT, size=10, color=INK); sc.alignment = LEFT; sc.border = BORDER
    ws0.row_dimensions[r].height = 20
    r += 1
for col, wd in zip("ABCDEF", [22, 16, 10, 10, 10, 30]):
    ws0.column_dimensions[col].width = wd
ws0.sheet_view.showGridLines = False

# ── Sheet2 测试用例 ───────────────────────────────────────────
ws = wb.create_sheet("测试用例")
title_block(ws, "测试用例 · 执行时填『结果/实测记录』两列", "共 %d 条(P0必过 5 · P0 %d · P1 %d · 暂缓 2)· 结果列有下拉" % (
    len(CASES),
    sum(1 for c in CASES if c[2] == "P0"),
    sum(1 for c in CASES if c[2] == "P1")), len(COLS))
hr = 4
for j, name in enumerate(COLS, 1):
    ws.cell(row=hr, column=j, value=name)
style_header(ws, hr, len(COLS))
for i, case in enumerate(CASES):
    rr = hr + 1 + i
    vals = list(case) + ["未测", "", ""]
    for j, val in enumerate(vals, 1):
        cell = ws.cell(row=rr, column=j, value=val)
        cell.font = Font(name=FONT, size=10, color=INK)
        cell.border = BORDER
        cell.alignment = CENTER if j in (1, 2, 3, 8, 9) else LEFT
    gf = GROUP_FILL.get(case[1])
    if gf: ws.cell(row=rr, column=2).fill = PatternFill("solid", fgColor=gf)
    pf = PRI_FILL.get(case[2])
    if pf: ws.cell(row=rr, column=3).fill = PatternFill("solid", fgColor=pf)
    if case[2] == "P0必过":
        ws.cell(row=rr, column=3).font = Font(name=FONT, size=10, bold=True, color="B91C1C")
    if case[2] == "暂缓":
        for j in range(1, len(COLS) + 1):
            ws.cell(row=rr, column=j).font = Font(name=FONT, size=10, color="9CA3AF")
        ws.cell(row=rr, column=9).value = "跳过"
    ws.row_dimensions[rr].height = 44
dv = DataValidation(type="list", formula1='"未测,通过,失败,阻塞,跳过"', allow_blank=False)
ws.add_data_validation(dv)
dv.add(f"I{hr+1}:I{hr+len(CASES)}")
for j, wd in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(j)].width = wd
ws.freeze_panes = "B5"
ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(COLS))}{hr+len(CASES)}"
ws.sheet_view.showGridLines = False

# ── Sheet3 设备准备 ───────────────────────────────────────────
ws3 = wb.create_sheet("设备准备")
title_block(ws3, "设备准备清单", "测试开始前逐项打勾;来源 DEVICES.md + PROGRESS 待办", 5)
PREP = [
    ("眼镜/包","Rokid Glasses 开启 AIUI 调试(Rokid AI APP→设置→开发者→AIUI 调试)","前置"),
    ("眼镜/包","Agent 重新打包上传:必须带 accelerometer 权限(否则 IMU 计步全挂,A-03 必失败)","🔴 硬前置"),
    ("眼镜/包","手机 Rokid AI APP 已登录,可运行当前 Agent;手机网络可用","前置"),
    ("ESP32 模拟器","烧录 tools/esp32_hr_sim/esp32_hr_sim.ino(标准 HRS 0x180D,60-170bpm,断开自动重广播)","测 A-05/06/07/H-01/H-06"),
    ("Garmin Fenix 8","表上开:设置→传感器与配件→手腕式心率→广播心率(只用标准 HR 广播,不用 ANT+)","测 A-05/06/07 + B-03 对照"),
    ("Apple Watch S7","无需任何配置,反面用例(验证非标设备被忽略)","测 H-04"),
    ("安卓备用机","装 nRF Connect,建 GATT server 模拟 HR=0/残包/notify 停","测 H-03(可选)"),
    ("Chronos 手表","出厂私有协议未重刷 → 本轮不用","跳过"),
    ("场地","400m 操场或已知 1km 直路(B-03 距离标定);手机可开飞行模式(A-10)","外场批次"),
]
hr3 = 4
for j, name in enumerate(["设备/项", "准备步骤", "用途", "完成"], 1):
    ws3.cell(row=hr3, column=j, value=name)
style_header(ws3, hr3, 4)
for i, (dev, step, use) in enumerate(PREP):
    rr = hr3 + 1 + i
    for j, val in enumerate([dev, step, use, ""], 1):
        cell = ws3.cell(row=rr, column=j, value=val)
        cell.font = Font(name=FONT, size=10, color=INK)
        cell.border = BORDER
        cell.alignment = LEFT if j == 2 else CENTER
    if "硬前置" in use:
        ws3.cell(row=rr, column=3).font = Font(name=FONT, size=10, bold=True, color="B91C1C")
    ws3.row_dimensions[rr].height = 30
dv3 = DataValidation(type="list", formula1='"✓,✗,—"', allow_blank=True)
ws3.add_data_validation(dv3)
dv3.add(f"D{hr3+1}:D{hr3+len(PREP)}")
for j, wd in enumerate([16, 52, 22, 7], 1):
    ws3.column_dimensions[get_column_letter(j)].width = wd
ws3.sheet_view.showGridLines = False

# ── Sheet4 本轮口径 ───────────────────────────────────────────
ws4 = wb.create_sheet("本轮口径")
title_block(ws4, "本轮口径与暂缓项", "决策记录(2026-07-08),避免执行中翻旧文档", 3)
SCOPE = [
    ("做", "硬件主线:IMU 计步/BLE 心率(连接·断连·重连·边界包)/息屏恢复/物理按键/语音闭环(宿主 ASR/LLM/TTS)/断网兜底"),
    ("做", "无 key 降级路径(A-12a/A-14a):验证『不登录一切正常』——不发登录请求、教练可用、摘要只入队不上传"),
    ("暂缓", "登录/后端链路(A-12b 记忆、A-14b 落库补传):眼镜端暂不做登录;coach_app_key 不注入;生产 nginx bug 修复不阻塞本轮"),
    ("暂缓", "账号绑定(匿名↔手机号)、跑后总结卡、自定义 maxHr —— 均列 M1,不阻塞硬件验证"),
    ("提醒", "若后续开通后端链路:先注入 coach_app_key、修生产 nginx POST 转发 bug,再补测 A-12b/A-14b 两条即可,用例已写好"),
]
r4 = 4
for tag, txt in SCOPE:
    tc = ws4.cell(row=r4, column=1, value=tag)
    tc.font = Font(name=FONT, bold=True, size=10, color=("0F5132" if tag == "做" else ("B91C1C" if tag == "暂缓" else "92400E")))
    tc.alignment = CENTER; tc.border = BORDER
    tc.fill = PatternFill("solid", fgColor=("D1FAE5" if tag == "做" else ("FEE2E2" if tag == "暂缓" else "FEF3C7")))
    ws4.merge_cells(start_row=r4, start_column=2, end_row=r4, end_column=3)
    xc = ws4.cell(row=r4, column=2, value=txt)
    xc.font = Font(name=FONT, size=10, color=INK); xc.alignment = LEFT; xc.border = BORDER
    ws4.row_dimensions[r4].height = 34
    r4 += 1
for j, wd in enumerate([8, 60, 30], 1):
    ws4.column_dimensions[get_column_letter(j)].width = wd
ws4.sheet_view.showGridLines = False

out = "/Users/apple/Desktop/Project/AIUI_AISmartRun/docs/AIUI眼镜端硬件测试执行表.xlsx"
wb.save(out)
print("saved:", out, "| cases:", len(CASES))
